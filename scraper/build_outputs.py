"""Rank teams, run sanity checks, and write the dashboard data + Excel workbook.

Reads ``scraper/_raw_data.json`` (produced by scrape.py) and writes into
``docs/data/``:
  - season.json, last7.json   team totals + ranks per period
  - meta.json                 category metadata + last-updated timestamp
  - mlb_team_stats.xlsx       downloadable workbook (one sheet per period)

    python scraper/build_outputs.py
"""

from __future__ import annotations

import datetime as dt
import json
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scrape import CATEGORIES, scrape_all

HERE = os.path.dirname(__file__)
RAW_PATH = os.path.join(HERE, "_raw_data.json")
DATA_DIR = os.path.join(HERE, "..", "docs", "data")

NUM_TEAMS = 30
PERIOD_LABELS = {"season": "Season", "last7": "Last 7 Days"}

# Plausible value ranges for sanity checks, keyed by category key.
RANGES = {
    "avg": (0, 1), "obp": (0, 1), "slg": (0, 1.2), "baa": (0, 1),
    "era": (0, 20), "whip": (0, 4),
    "so": (0, 5000), "runs": (0, 5000), "hr": (0, 1000), "hr_against": (0, 1000),
}


def assign_ranks(teams: list[dict], key: str, better: str) -> None:
    """Competition ranking (shared *lower* rank on ties); rank 1 == best."""
    reverse = better == "high"
    ordered = sorted(teams, key=lambda t: t["stats"][key]["value"], reverse=reverse)
    last_val = object()
    last_rank = 0
    for i, team in enumerate(ordered):
        val = team["stats"][key]["value"]
        if val == last_val:
            rank = last_rank
        else:
            rank = i + 1
            last_rank, last_val = rank, val
        team["stats"][key]["rank"] = rank


def compute_ranks(data: dict) -> None:
    for teams in data["periods"].values():
        for key, _label, _group, _ec, _en, better, _fmt in CATEGORIES:
            assign_ranks(teams, key, better)


def sanity_check(data: dict) -> None:
    for period, teams in data["periods"].items():
        assert len(teams) == NUM_TEAMS, f"{period}: expected {NUM_TEAMS}, got {len(teams)}"
        for key, label, _g, _ec, _en, _better, _fmt in CATEGORIES:
            ranks = sorted(t["stats"][key]["rank"] for t in teams)
            assert ranks[0] == 1, f"{period}/{key}: min rank {ranks[0]} != 1"
            assert max(ranks) <= NUM_TEAMS, f"{period}/{key}: max rank {max(ranks)} > {NUM_TEAMS}"
            lo, hi = RANGES[key]
            for t in teams:
                v = t["stats"][key]["value"]
                assert v is not None, f"{period}/{key}: null value for {t['abbr']}"
                assert lo <= v <= hi, f"{period}/{key}: {t['abbr']} value {v} out of [{lo},{hi}]"
    print("sanity checks passed")


def category_meta() -> list[dict]:
    return [
        {"key": k, "label": label, "group": group, "better": better, "fmt": fmt}
        for k, label, group, _ec, _en, better, fmt in CATEGORIES
    ]


def write_json(data: dict) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    for period, teams in data["periods"].items():
        payload = {"period": period, "teams": teams}
        with open(os.path.join(DATA_DIR, f"{period}.json"), "w", encoding="utf-8") as fh:
            json.dump(payload, fh, separators=(",", ":"))
    meta = {
        "updated": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "seasonYear": data.get("seasonYear"),
        "categories": category_meta(),
    }
    with open(os.path.join(DATA_DIR, "meta.json"), "w", encoding="utf-8") as fh:
        json.dump(meta, fh, indent=2)


def write_excel(data: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2A37")
    center = Alignment(horizontal="center")

    for period in ("season", "last7"):
        teams = data["periods"][period]
        ws = wb.create_sheet(PERIOD_LABELS[period])

        # Header: Team, then <Label> and <Label> Rank per category.
        headers = ["Team"]
        for _k, label, *_ in CATEGORIES:
            headers.extend([label, f"{label} Rank"])
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        # Sort rows by team name for a stable, readable workbook.
        for team in sorted(teams, key=lambda t: t["name"]):
            row = [team["name"]]
            for key, *_ in [(c[0],) for c in CATEGORIES]:
                s = team["stats"][key]
                row.extend([s["value"], s["rank"]])
            ws.append(row)

        # Color-scale the rank columns: rank 1 (best) green -> 30 (worst) red.
        last_row = ws.max_row
        rule = ColorScaleRule(
            start_type="num", start_value=1, start_color="63BE7B",
            mid_type="num", mid_value=15, mid_color="FFEB84",
            end_type="num", end_value=NUM_TEAMS, end_color="F8696B",
        )
        for idx in range(len(CATEGORIES)):
            rank_col = 3 + idx * 2  # Team(1), then value,rank pairs
            letter = get_column_letter(rank_col)
            ws.conditional_formatting.add(f"{letter}2:{letter}{last_row}", rule)

        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 24

    out = os.path.join(DATA_DIR, "mlb_team_stats.xlsx")
    wb.save(out)
    print(f"wrote {out}")


def main() -> None:
    if os.path.exists(RAW_PATH):
        with open(RAW_PATH, encoding="utf-8") as fh:
            data = json.load(fh)
    else:  # allow running without a separate scrape step
        data = scrape_all()

    compute_ranks(data)
    sanity_check(data)
    write_json(data)
    write_excel(data)
    print("done")


if __name__ == "__main__":
    main()
